"""JSON to Excel conversion logic."""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from column_dict import SEGMENT_MAP, WORD_MAP


class ConversionError(Exception):
    """Raised when JSON cannot be converted to Excel."""


def _contains_chinese(text: str) -> bool:
    return bool(re.search(r"[\u4e00-\u9fff]", text))


def _segment_lookup_key(segment: str) -> str:
    normalized = re.sub(r"([a-z0-9])([A-Z])", r"\1_\2", segment)
    return re.sub(r"[_\-\s]+", "_", normalized).lower().strip("_")


def _split_key_part(part: str) -> list[str]:
    normalized = re.sub(r"([a-z0-9])([A-Z])", r"\1_\2", part)
    normalized = re.sub(r"([A-Z]+)([A-Z][a-z])", r"\1_\2", normalized)
    return [token for token in re.split(r"[_\-\s]+", normalized.lower()) if token]


def _translate_token(token: str) -> str:
    if token in WORD_MAP:
        return WORD_MAP[token]
    if token.endswith("s") and token[:-1] in WORD_MAP:
        return WORD_MAP[token[:-1]]
    if token.endswith("id") and len(token) > 2:
        prefix = WORD_MAP.get(token[:-2], token[:-2])
        return f"{prefix}编号" if prefix else "编号"
    return token


def _translate_segment(segment: str) -> str:
    if _contains_chinese(segment):
        return segment
    lookup_key = _segment_lookup_key(segment)
    if lookup_key in SEGMENT_MAP:
        return SEGMENT_MAP[lookup_key]
    compact_key = lookup_key.replace("_", "")
    if compact_key in SEGMENT_MAP:
        return SEGMENT_MAP[compact_key]
    tokens = _split_key_part(segment)
    if not tokens:
        return segment
    translated = [_translate_token(token) for token in tokens]
    result = "".join(part for part in translated if part)
    return result or segment


def translate_column_name(column: str, overrides: dict[str, str] | None = None) -> str:
    """将英文字段名翻译为中文含义；已是中文或无法识别时保留原样。"""
    column = str(column).strip()
    if overrides and column in overrides:
        return overrides[column]
    if _contains_chinese(column):
        return column
    parts = column.split(".")
    translated = [_translate_segment(part) for part in parts]
    if len(translated) == 1:
        return translated[0]
    return "·".join(translated)


def translate_columns(df: pd.DataFrame, overrides: dict[str, str] | None = None) -> pd.DataFrame:
    df = df.copy()
    df.columns = [translate_column_name(col, overrides) for col in df.columns]
    return df


def get_column_mappings(text: str, overrides: dict[str, str] | None = None) -> tuple[str, list[tuple[str, str, str]] | None]:
    """返回字段映射列表：(英文字段, 默认中文, 当前中文)。"""
    ok, message, data = validate_json(text)
    if not ok:
        return message, None
    try:
        df = _to_dataframe(data)
    except ConversionError as exc:
        return str(exc), None

    mappings: list[tuple[str, str, str]] = []
    for col in df.columns:
        original = str(col)
        default_cn = translate_column_name(original)
        current_cn = overrides.get(original, default_cn) if overrides else default_cn
        mappings.append((original, default_cn, current_cn))
    return f"{message} | 共 {len(mappings)} 个字段", mappings


def validate_json(text: str) -> tuple[bool, str, Any | None]:
    """Validate JSON text. Returns (ok, message, parsed_data)."""
    text = text.strip()
    if not text:
        return False, "JSON 内容为空", None
    try:
        data = json.loads(text)
    except json.JSONDecodeError as exc:
        return False, f"JSON 格式错误（第 {exc.lineno} 行，第 {exc.colno} 列）: {exc.msg}", None
    return True, "JSON 格式正确", data


def _to_dataframe(data: Any) -> pd.DataFrame:
    if isinstance(data, list):
        if not data:
            raise ConversionError("JSON 数组为空，无法生成表格")
        if all(isinstance(item, dict) for item in data):
            return pd.json_normalize(data)
        if all(not isinstance(item, (dict, list)) for item in data):
            return pd.DataFrame({"值": data})
        raise ConversionError("不支持的数组元素类型，请使用对象数组或简单值数组")

    if isinstance(data, dict):
        if all(isinstance(v, list) for v in data.values()):
            non_empty = [v for v in data.values() if v]
            if non_empty and all(isinstance(item, dict) for sub in non_empty for item in sub):
                frames = {str(k): pd.json_normalize(v) for k, v in data.items() if v}
                if len(frames) == 1:
                    return next(iter(frames.values()))
                raise ConversionError(
                    "检测到多个数组字段，当前版本请使用单一对象数组，或多 Sheet 功能后续支持"
                )
        return pd.json_normalize(data)

    return pd.DataFrame({"值": [data]})


def format_field_label(column: str, overrides: dict[str, str] | None = None) -> str:
    """生成字段下拉显示文本：英文名（中文含义）。"""
    column = str(column)
    translated = translate_column_name(column, overrides)
    if translated != column:
        return f"{column}（{translated}）"
    return column


def parse_field_label(label: str) -> str:
    """从下拉显示文本还原原始字段名。"""
    if "（" in label:
        return label.split("（", 1)[0]
    return label.strip()


def get_available_fields(text: str, overrides: dict[str, str] | None = None) -> tuple[str, list[str] | None]:
    """解析 JSON 并返回可用于拆分的字段列表（带中文标注的显示标签）。"""
    ok, message, data = validate_json(text)
    if not ok:
        return message, None
    try:
        df = _to_dataframe(data)
    except ConversionError as exc:
        return str(exc), None
    labels = [format_field_label(col, overrides) for col in df.columns]
    return f"{message} | 共 {len(labels)} 个字段", labels


def get_preview_info(text: str, overrides: dict[str, str] | None = None) -> tuple[str, pd.DataFrame | None]:
    ok, message, data = validate_json(text)
    if not ok:
        return message, None
    try:
        df = translate_columns(_to_dataframe(data), overrides)
    except ConversionError as exc:
        return str(exc), None
    return f"{message} | 共 {len(df)} 行 × {len(df.columns)} 列", df


def _group_key(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "__NULL__"
    return str(value)


def _display_group_value(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "（空值）"
    return str(value)


def _resolve_split_column(
    df_raw: pd.DataFrame,
    df: pd.DataFrame,
    split_by: str,
    overrides: dict[str, str] | None = None,
) -> str:
    """根据原始或翻译后的字段名，定位拆分列在翻译后 DataFrame 中的列名。"""
    split_by = split_by.strip()
    if not split_by:
        raise ConversionError("未指定拆分字段")

    original_columns = [str(col) for col in df_raw.columns]
    if split_by in original_columns:
        return str(df.columns[original_columns.index(split_by)])

    translated_columns = [str(col) for col in df.columns]
    if split_by in translated_columns:
        return split_by

    for idx, original in enumerate(original_columns):
        if translate_column_name(original, overrides) == split_by:
            return str(df.columns[idx])

    raise ConversionError(f"找不到拆分字段「{split_by}」，请先点击「解析字段」")


def get_split_preview(
    text: str,
    split_by: str,
    overrides: dict[str, str] | None = None,
) -> tuple[str, list[dict[str, Any]] | None]:
    """预览按字段拆分结果，供用户确认和自定义 Sheet 名。"""
    ok, message, data = validate_json(text)
    if not ok:
        return message, None
    try:
        df_raw = _to_dataframe(data)
        df = translate_columns(df_raw, overrides)
        split_column = _resolve_split_column(df_raw, df, split_by, overrides)
    except ConversionError as exc:
        return str(exc), None

    groups = sorted(df.groupby(split_column, sort=False), key=lambda item: str(item[0]))
    preview: list[dict[str, Any]] = []
    for group_value, group_df in groups:
        preview.append(
            {
                "value_key": _group_key(group_value),
                "display_value": _display_group_value(group_value),
                "row_count": len(group_df),
                "sheet_name": _sanitize_sheet_name(group_value, set()),
            }
        )
    return f"将拆分为 {len(preview)} 个工作表", preview


def _sanitize_sheet_name(name: Any, used: set[str]) -> str:
    if name is None or (isinstance(name, float) and pd.isna(name)):
        text = "空值"
    else:
        text = str(name).strip() or "空值"

    text = re.sub(r"[\\/*?:\[\]]", "_", text)
    if not text:
        text = "空值"

    base = text[:31]
    candidate = base
    counter = 2
    while candidate in used:
        suffix = f"_{counter}"
        candidate = f"{base[: 31 - len(suffix)]}{suffix}"
        counter += 1
    used.add(candidate)
    return candidate


def _style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_font = Font(bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for ws in wb.worksheets:
        if ws.max_row == 0:
            continue
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = header_alignment

        for col_idx in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for row_idx in range(1, ws.max_row + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is not None:
                    max_length = max(max_length, len(str(value)))
            ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 50)

        ws.freeze_panes = "A2"

    wb.save(path)


def convert_json_to_excel(
    text: str,
    output_path: Path,
    split_by: str | None = None,
    sheet_names: dict[str, str] | None = None,
    column_overrides: dict[str, str] | None = None,
) -> str:
    ok, message, data = validate_json(text)
    if not ok:
        raise ConversionError(message)

    try:
        df_raw = _to_dataframe(data)
        df = translate_columns(df_raw, column_overrides)
    except ConversionError:
        raise
    except Exception as exc:
        raise ConversionError(f"数据转换失败: {exc}") from exc

    output_path = Path(output_path)
    if output_path.suffix.lower() != ".xlsx":
        output_path = output_path.with_suffix(".xlsx")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    if split_by:
        split_column = _resolve_split_column(df_raw, df, split_by, column_overrides)
        used_names: set[str] = set()
        groups = sorted(df.groupby(split_column, sort=False), key=lambda item: str(item[0]))

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            for group_value, group_df in groups:
                key = _group_key(group_value)
                if sheet_names and key in sheet_names:
                    raw_name = sheet_names[key]
                else:
                    raw_name = _display_group_value(group_value)
                sheet_name = _sanitize_sheet_name(raw_name, used_names)
                group_df.to_excel(writer, sheet_name=sheet_name, index=False)

        _style_workbook(output_path)
        sheet_count = len(groups)
        return (
            f"已成功生成: {output_path.name}（{len(df)} 行 × {len(df.columns)} 列，"
            f"按「{translate_column_name(split_by, column_overrides)}」拆分为 {sheet_count} 个工作表）"
        )

    df.to_excel(output_path, index=False, engine="openpyxl")
    _style_workbook(output_path)

    return f"已成功生成: {output_path.name}（{len(df)} 行 × {len(df.columns)} 列）"
